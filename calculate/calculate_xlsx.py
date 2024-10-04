from flask import Flask, request, Response
import pandas as pd
from openpyxl import load_workbook
import io
import os
import shutil

app = Flask(__name__)

def process_dataframe_in_chunks(df, chunk_size=500):
    """Yield chunks of the DataFrame."""
    for start in range(0, df.shape[0], chunk_size):
        yield df.iloc[start:start + chunk_size]

def create_new_excel_with_data(df, template_path):
    """Create a new Excel file based on the template and update it with the DataFrame data."""
    output = io.BytesIO()
    
    # Create a temporary copy of the template
    temp_file_path = 'calculate/temp_formulat.xlsx'
    shutil.copy(template_path, temp_file_path)
    
    workbook = load_workbook(temp_file_path)
    sheet = workbook.active

    # Assuming your data starts from row 2 to keep headers intact.
    start_row = 2
    for chunk in process_dataframe_in_chunks(df):
        for i, row in chunk.iterrows():
            for j, value in enumerate(row):
                # Assuming columns start from A, B, C, ... (0-indexed)
                sheet.cell(row=start_row + i, column=j + 1, value=value)
                print(f'col: {j}, row: {start_row + i}, value: {value}' )
                # print('col: ', j)
                # print('value: ', value)
                # print('\n')

    workbook.save(output)
    output.seek(0)
    
    # Clean up the temporary file
    os.remove(temp_file_path)
    
    return output



def run(df):
    formulat_file_path = 'calculate/formula.xlsx' 
    # Read the existing Excel file
    c = create_new_excel_with_data(df, formulat_file_path)
    # Stream the new Excel file
    return Response(c, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment;filename=output_with_calculations.xlsx"})

