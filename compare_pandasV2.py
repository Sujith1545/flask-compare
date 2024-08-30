import openpyxl
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill
# from io import BytesIO
import os
from io import StringIO, BytesIO
# from Utilities import constant_config as constant
# constant = {
#     UNIQUE_QUEUES: ["GEO","PROGRAM"]
# }

# print(constant.UNIQUE_QUEUES)

pd.options.mode.chained_assignment = None


# def get_headers(f):
#     df = pd.read_excel(f, engine='openpyxl')
#     return list(df.columns.values)

# def get_headersV2(df):
#     df1 = df.loc[:, ~df.columns.str.contains('^Unnamed')]
#     h = list(df1.columns.values)
#     return h

# def read_file(f):
#     split_tup = os.path.splitext(f.filename)
#     file_extension = split_tup[1]
#     if file_extension == '.csv':
#         df = pd.read_csv(f)
#         df.drop(df.columns[df.columns.str.contains('unnamed',case = False)],axis = 1, inplace = True)
#         # df1 = df.loc[:, ~df.columns.str.contains('^Unnamed')]
#         return df
#     if file_extension == '.xlsx':
#         df = pd.read_excel(f, engine='openpyxl')
#         # df1 = df.loc[:, ~df.columns.str.contains('^Unnamed')]
#         df.drop(df.columns[df.columns.str.contains('unnamed',case = False)],axis = 1, inplace = True)
#         return df
#     return None


# def to_xlsx(in_file_path, out_file_path):
#     split_tup = os.path.splitext(in_file_path)
#     file_extension = split_tup[1]
#     if file_extension == '.csv':
#         read_file = pd.read_csv(in_file_path)
#         read_file.to_excel(out_file_path, index=None, header=True)
#     if file_extension == '.xlsx':
#         df = pd.read_excel(in_file_path, engine='openpyxl')
#         df.to_excel(out_file_path)
#     return
    

# def get_df_v2(fObj):
#     split_tup = os.path.splitext(fObj.filename)
#     file_extension = split_tup[1]
#     if file_extension == '.xlsx':
#         df = pd.read_excel(fObj, engine='openpyxl')
#         return df
#     if file_extension == '.csv':
#         print('fObj: ', fObj)
#         df = pd.read_csv('uploads/file1.xlsx', sep="\s+")
#         return df
#     return None

# def get_df(f):
#     # split_tup = os.path.splitext('my_file.txt')
#     df = pd.read_excel(f, engine='openpyxl')
#     return df

def compare_headers(df2, selected_headers):
    # Find headers selected in file 1 that are missing in file 2
    missing_headers = [header for header in selected_headers if header not in df2.columns]

    if missing_headers:
        return f"Missing headers in File 2: {', '.join(missing_headers)}"
    return None

"""
    df1 : Dataframe
    df2 : Data frame
    file_path_df_result: store file path of comparison
    primary_key : unique based need to filter and compare similar like primary key  
                    constant.UNIQUE_QUEUES = ["GEO","PROGRAM"]
    is_tolerance : False default for no need to check tolerance limit.
                    True enabled means it will consider for tolerance 
"""
def pandas_compare(df1, df2, primary_key, is_tolerance=False):
    print('primary_key:: ', primary_key)
    """Function will compare two dataframes even if rows count are not matching and stores the original two files,
       additional rows from both dataframes if present and the delta in an Excel file."""
    df1.replace(np.nan, 0, inplace=True)
    df2.replace(np.nan, 0, inplace=True)

    # print('df1: ', df1)
    # print('df2: ', df2)

    # Merge TWO DF - if all values match - both , if miss match  left  df1 and right  df2 will show
    df_mismatch = pd.merge(df1, df2, how='outer', indicator=True)

    # print('df_mismatch: ', df_mismatch)
    # in below two condition ignoring both value and getting additional data for comparing
    # Selecting the rows which are extra(additional data) in df1 and mismatch rows present in df1
    df_mismatch1 = df_mismatch[df_mismatch['_merge'] == 'left_only']
    df_mismatch1 = df_mismatch1.drop(['_merge'], axis=1)
    # Selecting the rows which are extra(additional data) in df2 and mismatch rows present in df2
    df_mismatch2 = df_mismatch[df_mismatch['_merge'] == 'right_only']
    df_mismatch2 = df_mismatch2.drop(['_merge'], axis=1)


    miss_match_df = pd.merge(df_mismatch1, df_mismatch2, how='outer', on=primary_key, indicator=True)

    if len(miss_match_df[miss_match_df['_merge'] == 'both']) == 0 and (len(miss_match_df[miss_match_df['_merge']
                                                                                         == 'left_only']) > 0 or len(
        miss_match_df[miss_match_df['_merge'] == 'right_only']) > 0):
        add_df1 = df_mismatch1
        add_df2 = df_mismatch2
        delta_mismatch = pd.DataFrame()
    # Filtering the mismatch values and additional values in different dataframes
    else:

        filter_unique_key = "|".join(primary_key)

        # Filtering out the mismatch values(not additional) in both dataframes
        df_mismatch1, df_mismatch2 = parse_data_df_using_regex(miss_match_df, filter_unique_key, method_type='both')
        
        df_mismatch1.insert(0, 'type', 'file1')
        df_mismatch2.insert(0, 'type', 'file2')

        print('--df_mismatch1--: ',  df_mismatch1)
        print('--df_mismatch2--: ',  df_mismatch2)


        # concatenating mismatch data from first dataframe row wise
        delta_mismatch = pd.concat([df_mismatch1, df_mismatch2]).sort_index(kind='merge')
        # print('delta_mismatch: ', delta_mismatch)

        # Now checking Additional rows from dataframe 1 & dataframe 2
        add_df1, add_df2 = parse_data_df_using_regex(miss_match_df, filter_unique_key,
                                                     method_type='left_only|right_only')

    if len(delta_mismatch) > 0 or len(add_df1) > 0 or len(add_df2) > 0:
        print("MISMATCH FOUND.........................")
        result_stream = BytesIO()
        try:
            with pd.ExcelWriter(result_stream, engine="openpyxl") as writer:
                add_df1.to_excel(writer, sheet_name='Actual_Additional', index=False)
                add_df2.to_excel(writer, sheet_name='Expected_Additional', index=False)
                delta_mismatch.to_excel(writer, sheet_name='Mismatch', index=False) 
             
        except Exception as e:
            print("Not able to create file11 {}")
            print(e)

        column_length = len(delta_mismatch.columns)
        if len(delta_mismatch) > 0:
            print("Highlighting started for mismatches")
            add_background_colors(file_path_df_result, column_length)
        
        result_stream.seek(0) 
        return result_stream 
    else:
        print("NO MISMATCH FOUND.........................")
        return None


def parse_data_df_using_regex(data_frame, regex_value, method_type='both'):
    # To remove _x and _y value and slice by x and y values to filter to create two df
    if '|' in method_type:
        method_type_value = method_type.split('|')[0]
    else:
        method_type_value = method_type
    common_df = data_frame[data_frame['_merge'] == method_type_value]
    common_df = common_df.drop(['_merge'], axis=1)
    df1 = common_df.filter(regex=str(regex_value) + '|_x')
    df1.columns = df1.columns.str.replace('_x', '')

    if '|' in method_type:
        method_type_value = method_type.split('|')[1]
        common_df = data_frame[data_frame['_merge'] == method_type_value]
        common_df = common_df.drop(['_merge'], axis=1)

    df2 = common_df.filter(regex=str(regex_value) + '|_y')
    df2.columns = df2.columns.str.replace('_y', '')

    return df1, df2


def add_background_colors(path, column_length):
    wb = openpyxl.load_workbook(path)
    ws = wb['Mismatch']
    fill_cell = PatternFill(patternType='solid', fgColor='FC2C03')
    for i in range(2, ws.max_row, 2):
        for j in range(6, column_length + 1):
            if ws.cell(row=i, column=j).value == ws.cell(row=i + 1, column=j).value:
                continue
            else:
                ws.cell(row=i, column=j).fill = fill_cell
                ws.cell(row=i + 1, column=j).fill = fill_cell
    wb.save(path)


# if __name__ == '__main__':
#     pass